from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


CONTENT_TYPE_XLSX = (
    'application/vnd.openxmlformats-officedocument.'
    'spreadsheetml.sheet'
)


def ajustar_larguras_colunas(
    worksheet,
    *,
    largura_maxima=50,
    margem=4,
):
    for coluna in worksheet.columns:
        valores = [
            str(celula.value)
            for celula in coluna
            if celula.value is not None
        ]

        maior = max(
            (len(valor) for valor in valores),
            default=0,
        )

        largura = min(
            maior + margem,
            largura_maxima,
        )

        letra = coluna[0].column_letter

        worksheet.column_dimensions[
            letra
        ].width = max(
            largura,
            10,
        )


def estilizar_cabecalho(worksheet):
    preenchimento = PatternFill(
        fill_type='solid',
        fgColor='1F2937',
    )

    fonte = Font(
        color='FFFFFF',
        bold=True,
    )

    for celula in worksheet[1]:
        celula.fill = preenchimento
        celula.font = fonte
        celula.alignment = Alignment(
            horizontal='center',
            vertical='center',
        )


def gerar_modelo_excel(
    *,
    nome_aba,
    colunas,
    linhas_exemplo=None,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = nome_aba[:31]

    worksheet.append(
        list(colunas)
    )

    for linha in linhas_exemplo or []:
        worksheet.append(
            list(linha)
        )

    worksheet.freeze_panes = 'A2'

    estilizar_cabecalho(
        worksheet
    )

    ajustar_larguras_colunas(
        worksheet
    )

    arquivo = BytesIO()
    workbook.save(
        arquivo
    )
    arquivo.seek(0)

    return arquivo


def criar_resposta_download_excel(
    *,
    arquivo,
    nome_arquivo,
):
    response = HttpResponse(
        arquivo.getvalue(),
        content_type=CONTENT_TYPE_XLSX,
    )

    response['Content-Disposition'] = (
        f'attachment; filename="{nome_arquivo}"'
    )

    return response
